#!/usr/bin/env python3
"""Prepare groundwater nitrate station-year tables from the UBA Nitratbericht XLSX."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = (
    ROOT
    / "data"
    / "raw"
    / "uba_nitrate_report_2024"
    / "692b5285-78ca-4081-aa31-4b327cf0105f_XLSX"
    / "Download_Daten_Nitratbericht_2024.xlsx"
)
OUT_DIR = ROOT / "data" / "processed" / "uba_nitrate_report_2024"


STATION_COLUMNS = {
    "Messstellencode": "station_id",
    "Messstellenname": "station_name",
    "Wasserkörpercode": "water_body_id",
    "Wasserkörpername": "water_body_name",
    "Probenahmetiefe in m": "sampling_depth_m",
    "Länge in Grad (ETRS89)": "lon_etrs89",
    "Breite in Grad (ETRS89)": "lat_etrs89",
}

NITRATE_COLUMNS = {
    "Messstellencode": "station_id",
    "Jahr": "year",
    "Anzahl Messungen": "n_measurements",
    "Jahresmittelwert in mg Nitrat pro Liter": "nitrate_mg_l",
}


def normalise(value: object) -> object:
    if value is None:
        return ""
    return value


def read_sheet(path: Path, sheet_name: str, columns: dict[str, str]) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)
    selected = []
    names = []
    for index, name in enumerate(header):
        if name in columns:
            selected.append(index)
            names.append(columns[str(name)])

    missing = sorted(set(columns) - {str(name) for name in header if name is not None})
    if missing:
        raise ValueError(f"{sheet_name} is missing expected columns: {missing}")

    records = []
    for row in rows:
        record = {name: normalise(row[index]) for name, index in zip(names, selected)}
        if any(value != "" for value in record.values()):
            records.append(record)
    return records


def add_state(records: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    out = []
    for record in records:
        station_id = str(record["station_id"])
        state = station_id.split("_", 1)[0] if "_" in station_id else station_id[:2]
        out.append(record | {"state_code": state})
    return out


def index_by_station(records: Iterable[dict[str, object]]) -> dict[str, dict[str, object]]:
    return {str(record["station_id"]): record for record in records}


def merge_panel(
    nitrate_rows: Iterable[dict[str, object]],
    station_rows: Iterable[dict[str, object]],
) -> list[dict[str, object]]:
    stations = index_by_station(station_rows)
    panel = []
    for nitrate in nitrate_rows:
        station = stations.get(str(nitrate["station_id"]), {})
        panel.append(
            nitrate
            | {
                "state_code": station.get("state_code", ""),
                "station_name": station.get("station_name", ""),
                "water_body_id": station.get("water_body_id", ""),
                "water_body_name": station.get("water_body_name", ""),
                "sampling_depth_m": station.get("sampling_depth_m", ""),
                "lon_etrs89": station.get("lon_etrs89", ""),
                "lat_etrs89": station.get("lat_etrs89", ""),
            }
        )
    return panel


def write_csv(path: Path, records: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not records:
        raise ValueError(f"No records to write for {path}")
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(records[0]))
        writer.writeheader()
        writer.writerows(records)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise FileNotFoundError(f"UBA workbook not found: {args.xlsx}")

    stations = add_state(read_sheet(args.xlsx, "GW_Messstellen", STATION_COLUMNS))
    nitrate = read_sheet(args.xlsx, "GW_Nitrat_MW", NITRATE_COLUMNS)
    panel = merge_panel(nitrate, stations)
    panel_ni = [row for row in panel if row["state_code"] == "NI"]

    write_csv(args.out_dir / "gw_stations_de.csv", stations)
    write_csv(args.out_dir / "gw_nitrate_panel_de.csv", panel)
    write_csv(args.out_dir / "gw_nitrate_panel_ni.csv", panel_ni)

    summary = {
        "source_workbook": str(args.xlsx),
        "groundwater_stations_de": len(stations),
        "groundwater_panel_rows_de": len(panel),
        "groundwater_stations_ni": len({row["station_id"] for row in panel_ni}),
        "groundwater_panel_rows_ni": len(panel_ni),
        "year_min": min(int(row["year"]) for row in panel),
        "year_max": max(int(row["year"]) for row in panel),
        "outputs": [
            str(args.out_dir / "gw_stations_de.csv"),
            str(args.out_dir / "gw_nitrate_panel_de.csv"),
            str(args.out_dir / "gw_nitrate_panel_ni.csv"),
        ],
    }
    args.out_dir.mkdir(parents=True, exist_ok=True)
    (args.out_dir / "summary.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
