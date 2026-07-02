#!/usr/bin/env python3
"""Create the processed Niedersachsen nitrate station-year table."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
DEFAULT_INPUT = (
    ROOT
    / "data"
    / "uba_nitrate_report_2024"
    / "raw"
    / "Download_Daten_Nitratbericht_2024.xlsx"
)
DEFAULT_OUTPUT = (
    ROOT
    / "data"
    / "uba_nitrate_report_2024"
    / "processed"
    / "uba_nitrate_station_year_ni.csv"
)

STATION_COLUMNS = {
    "Messstellencode": "station_id",
    "Messstellenname": "station_name",
    "Wasserkörpercode": "water_body_id",
    "Wasserkörpername": "water_body_name",
    "Probenahmetiefe in m": "sampling_depth_m",
    "Länge in Grad (ETRS89)": "lon_etrs89",
    "Breite in Grad (ETRS89)": "lat_etrs89",
}
NITRATE_COLUMNS = {
    "Messstellencode": "station_id",
    "Jahr": "year",
    "Anzahl Messungen": "n_measurements",
    "Jahresmittelwert in mg Nitrat pro Liter": "nitrate_mg_l",
}


def normalise(value: object) -> object:
    return "" if value is None else value


def read_sheet(
    path: Path,
    sheet_name: str,
    columns: dict[str, str],
) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)
    indexes = [
        (index, columns[str(name)])
        for index, name in enumerate(header)
        if name in columns
    ]
    missing = sorted(set(columns) - {str(name) for name in header if name is not None})
    if missing:
        raise ValueError(f"{sheet_name} is missing columns: {missing}")

    records = []
    for row in rows:
        record = {name: normalise(row[index]) for index, name in indexes}
        if any(value != "" for value in record.values()):
            records.append(record)
    return records


def index_stations(
    records: Iterable[dict[str, object]],
) -> dict[str, dict[str, object]]:
    return {str(record["station_id"]): record for record in records}


def build_ni_table(
    station_rows: Iterable[dict[str, object]],
    nitrate_rows: Iterable[dict[str, object]],
) -> list[dict[str, object]]:
    stations = index_stations(station_rows)
    output = []
    for nitrate in nitrate_rows:
        station_id = str(nitrate["station_id"])
        if not station_id.startswith("NI_"):
            continue
        station = stations.get(station_id)
        if station is None:
            raise ValueError(f"Missing station metadata for {station_id}")
        output.append(
            nitrate
            | {
                "state_code": "NI",
                "station_name": station["station_name"],
                "water_body_id": station["water_body_id"],
                "water_body_name": station["water_body_name"],
                "sampling_depth_m": station["sampling_depth_m"],
                "lon_etrs89": station["lon_etrs89"],
                "lat_etrs89": station["lat_etrs89"],
            }
        )
    output.sort(key=lambda row: (str(row["station_id"]), int(row["year"])))
    return output


def write_csv(path: Path, records: list[dict[str, object]]) -> None:
    if not records:
        raise ValueError("UBA preprocessing produced no records")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(records[0]))
        writer.writeheader()
        writer.writerows(records)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.input.exists():
        raise FileNotFoundError(f"UBA workbook not found: {args.input}")

    stations = read_sheet(args.input, "GW_Messstellen", STATION_COLUMNS)
    nitrate = read_sheet(args.input, "GW_Nitrat_MW", NITRATE_COLUMNS)
    records = build_ni_table(stations, nitrate)
    write_csv(args.output, records)

    station_count = len({str(row["station_id"]) for row in records})
    years = [int(row["year"]) for row in records]
    print(f"Wrote {len(records)} rows for {station_count} stations")
    print(f"Years: {min(years)}-{max(years)}")
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
